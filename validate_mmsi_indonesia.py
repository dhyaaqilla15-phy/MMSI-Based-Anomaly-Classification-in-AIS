from __future__ import annotations

"""
Validasi MMSI terhadap aturan penggunaan di Indonesia dan menyortir entri tidak valid.

Contoh penggunaan:
    python validate_mmsi_indonesia.py --input sar_vessel_detections_pipev3_20250922.csv \
        --output sar_vessel_detections_mmsi_validation.xlsx
"""

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import PatternFill


MID_MIN = 201
MID_MAX = 775
VALID_FIRST_DIGITS = {"2", "3", "4", "5", "6", "7"}


@dataclass
class ValidationResult:
    is_valid: bool
    category: str
    note: str


def _extract_mid(mmsi: str, start: int) -> Optional[int]:
    """Return the MID (3 digits) starting at index `start`, or None if invalid."""
    end = start + 3
    if end > len(mmsi):
        return None
    segment = mmsi[start:end]
    if not segment.isdigit():
        return None
    return int(segment)


def _mid_in_range(mid: Optional[int]) -> bool:
    return mid is not None and MID_MIN <= mid <= MID_MAX


def validate_mmsi(mmsi: str) -> ValidationResult:
    """
    Validate MMSI according to Indonesian usage rules.

    Rules summary:
    - Ship station: digits 1-3 form MID (201-775) and first digit 2-7.
    - Handheld DSC: 8 + MID + 5 digits.
    - Free-form prefix 9: only 970/972/974 prefixes or 98/99 + MID + XXXX.
    - Group/Coast: start with 0 (group) or 00 (coast), MID still 201-775.
    """
    original = mmsi
    if mmsi is None:
        return ValidationResult(False, "unknown", "MMSI kosong")

    mmsi = str(mmsi).strip()
    if mmsi == "":
        return ValidationResult(False, "unknown", "MMSI kosong")

    if not mmsi.isdigit():
        return ValidationResult(False, "unknown", f"Mengandung karakter non-digit: {original}")

    if len(mmsi) != 9:
        return ValidationResult(False, "unknown", f"Panjang bukan 9 digit: {mmsi}")

    first_digit = mmsi[0]

    # Ship stations (2-7)
    if first_digit in VALID_FIRST_DIGITS:
        mid = _extract_mid(mmsi, 0)
        if _mid_in_range(mid):
            return ValidationResult(True, "ship_station", f"MID {mid}")
        return ValidationResult(False, "ship_station", f"MID {mid} di luar 201-775")

    # Handheld DSC starts with 8
    if first_digit == "8":
        mid = _extract_mid(mmsi, 1)
        if _mid_in_range(mid):
            return ValidationResult(True, "handheld_vhf_dsc", f"Prefix 8 dengan MID {mid}")
        return ValidationResult(False, "handheld_vhf_dsc", f"Setelah 8 bukan MID 201-775 (MID={mid})")

    # Free-form prefix 9
    if first_digit == "9":
        if mmsi.startswith("970"):
            return ValidationResult(True, "ais_sart", "970xxxxxx (AIS-SART)")
        if mmsi.startswith("972"):
            return ValidationResult(True, "mob_msld", "972xxxxxx (MOB/MSLD)")
        if mmsi.startswith("974"):
            return ValidationResult(True, "epirb_ais", "974xxxxxx (EPIRB-AIS)")
        if mmsi.startswith(("98", "99")):
            mid = _extract_mid(mmsi, 2)
            if _mid_in_range(mid):
                return ValidationResult(True, "auxiliary_craft", f"{mmsi[:2]} + MID {mid}")
            return ValidationResult(False, "auxiliary_craft", f"MID {mid} di luar 201-775")
        return ValidationResult(False, "free_form", "Prefix '9' tetapi bukan 970/972/974 atau 98/99 + MID")

    # Group or coast stations start with 0
    if first_digit == "0":
        if mmsi.startswith("00"):
            mid = _extract_mid(mmsi, 2)
            if _mid_in_range(mid):
                return ValidationResult(True, "coast_station", f"00 + MID {mid}")
            return ValidationResult(False, "coast_station", f"MID {mid} di luar 201-775")
        mid = _extract_mid(mmsi, 1)
        if _mid_in_range(mid):
            return ValidationResult(True, "group_call", f"0 + MID {mid}")
        return ValidationResult(False, "group_call", f"MID {mid} di luar 201-775")

    return ValidationResult(False, "unknown", f"Prefix {first_digit} tidak dikenal")


def analyze_csv(input_path: Path, output_path: Path) -> pd.DataFrame:
    """Read the CSV, validate MMSI entries, and write an Excel report with highlights."""
    df = pd.read_csv(input_path, dtype={"mmsi": "string"})
    df["mmsi"] = df["mmsi"].fillna("").str.strip()

    results = df["mmsi"].apply(validate_mmsi)
    df["mmsi_valid"] = results.map(lambda r: r.is_valid)
    df["mmsi_category"] = results.map(lambda r: r.category)
    df["mmsi_note"] = results.map(lambda r: r.note)

    sheet_name = "MMSI Validation"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        sheet = writer.sheets[sheet_name]
        invalid_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row_idx, is_valid in enumerate(df["mmsi_valid"], start=2):
            if not is_valid:
                for col_idx in range(1, len(df.columns) + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = invalid_fill

    return df


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Validasi MMSI sesuai aturan penggunaan di Indonesia dan sorot entri yang tidak valid."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("sar_vessel_detections_pipev3_20250922.csv"),
        help="Path ke file CSV sumber.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("sar_vessel_detections_mmsi_validation.xlsx"),
        help="Path file Excel hasil validasi.",
    )
    args = parser.parse_args()

    if not args.input.exists():
        raise FileNotFoundError(f"File input tidak ditemukan: {args.input}")

    df = analyze_csv(args.input, args.output)

    total = len(df)
    invalid_count = (~df["mmsi_valid"]).sum()
    print(f"Total entri: {total}")
    print(f"Tidak valid: {invalid_count}")
    if invalid_count:
        print("Contoh entri tidak valid:")
        print(df.loc[~df["mmsi_valid"], ["mmsi", "mmsi_note"]].head())

if __name__ == "__main__":
    main()
